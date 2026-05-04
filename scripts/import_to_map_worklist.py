"""
Import a filled "To Map (Active + In Stock)" sheet from
product_platform_overview_*.xlsx and apply the mappings.

Each filled row writes to TWO places (so the mapping survives the next
platform-CSV re-import which wipes platform_skus):
  1. platform_skus.internal_product_id + qty_per_sale  (immediate effect)
  2. ecommerce_listings (insert if not present, otherwise update)
     with a deterministic listing_key so re-running is idempotent.

Run: python scripts/import_to_map_worklist.py [PATH] [--dry-run]
     If PATH omitted, defaults to today's overview file.
"""
import argparse
import datetime
import hashlib
import sqlite3
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / 'inventory_app' / 'instance' / 'inventory.db'
EXPORTS = ROOT / 'data' / 'exports'
DEFAULT = EXPORTS / f'product_platform_overview_{datetime.date.today().strftime("%Y%m%d")}.xlsx'

EXPECTED_HEADERS = [
    'Platform', 'platform_sku_id', 'Parent ID', 'Variation ID', 'Listing name',
    'Variation', 'Seller SKU', 'Stock', 'Suggested SKU', 'Suggested name',
    'Confidence', '→ internal_sku (fill in)', '→ qty_per_sale (default 1)',
]


def listing_key_for(platform, variation_id, item_name, variation_name):
    """Deterministic key for ecommerce_listings.listing_key — survives re-runs."""
    base = f"{platform}|{variation_id or ''}|{item_name or ''}|{variation_name or ''}"
    return 'manual_' + hashlib.md5(base.encode('utf-8')).hexdigest()[:12]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('path', nargs='?', default=str(DEFAULT))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    if not Path(args.path).exists():
        sys.exit(f'File not found: {args.path}')

    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.startswith('To Map')), None)
    if not sheet_name:
        sys.exit('Worksheet "To Map (Active + In Stock)" not found in workbook')
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit('Empty worksheet')

    headers = list(rows[0])
    if headers[:len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        sys.exit(f'Header mismatch.\nExpected: {EXPECTED_HEADERS}\nGot:      {headers}')

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')

    stats = dict(scanned=0, applied=0, skipped_empty=0, sku_not_found=0,
                 ps_updated=0, listing_inserted=0, listing_updated=0)

    for raw in rows[1:]:
        if not raw or all(v in (None, '') for v in raw):
            continue
        stats['scanned'] += 1

        (platform, ps_id, parent_id, variation_id, item_name, variation_name,
         seller_sku, stock, sug_sku, sug_name, conf, fill_sku, fill_qty) = raw[:13]

        if fill_sku in (None, '') or str(fill_sku).strip() == '':
            stats['skipped_empty'] += 1
            continue

        try:
            sku_int = int(str(fill_sku).strip())
        except ValueError:
            print(f'  bad SKU "{fill_sku}" on platform_sku_id={ps_id} — skipped')
            stats['sku_not_found'] += 1
            continue

        prod = conn.execute(
            'SELECT id FROM products WHERE sku = ?', (sku_int,)
        ).fetchone()
        if not prod:
            print(f'  SKU {sku_int} not in products — platform_sku_id={ps_id} skipped')
            stats['sku_not_found'] += 1
            continue
        product_id = prod['id']

        try:
            qps = float(fill_qty) if fill_qty not in (None, '') else 1.0
            if qps <= 0:
                qps = 1.0
        except ValueError:
            qps = 1.0

        # 1) Update platform_skus directly (immediate effect)
        cur = conn.execute(
            '''UPDATE platform_skus
               SET internal_product_id = ?, qty_per_sale = ?
               WHERE id = ?''',
            (product_id, qps, ps_id)
        )
        if cur.rowcount:
            stats['ps_updated'] += 1

        # 2) Upsert ecommerce_listings so the next platform-CSV re-import
        #    auto-propagates this mapping back. Deterministic listing_key so
        #    re-running with the same xlsx is idempotent.
        lkey = listing_key_for(platform, variation_id, item_name, variation_name)
        existing = conn.execute(
            'SELECT id FROM ecommerce_listings WHERE listing_key = ?', (lkey,)
        ).fetchone()
        if existing:
            conn.execute(
                '''UPDATE ecommerce_listings
                   SET product_id = ?, qty_per_sale = ?,
                       item_name = ?, variation = ?, seller_sku = ?
                   WHERE id = ?''',
                (product_id, qps, item_name, variation_name, seller_sku, existing['id'])
            )
            stats['listing_updated'] += 1
        else:
            conn.execute(
                '''INSERT INTO ecommerce_listings
                   (platform, item_name, variation, seller_sku, listing_key,
                    sample_price, product_id, qty_per_sale)
                   VALUES (?, ?, ?, ?, ?, NULL, ?, ?)''',
                (platform, item_name, variation_name, seller_sku, lkey,
                 product_id, qps)
            )
            stats['listing_inserted'] += 1

        stats['applied'] += 1

    if args.dry_run:
        conn.rollback()
        print('DRY RUN — rolled back')
    else:
        conn.commit()
        print('Committed')
    conn.close()

    print()
    print('Summary:')
    for k, v in stats.items():
        print(f'  {k:22s} {v}')


if __name__ == '__main__':
    main()
