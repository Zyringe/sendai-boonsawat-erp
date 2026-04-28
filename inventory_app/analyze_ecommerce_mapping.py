"""
analyze_ecommerce_mapping.py — One-time analysis script
Reads all Shopee/Lazada order files, suggests ERP mapping, cross-checks vs BSN sales.
Output: data/exports/ecommerce_mapping_analysis_<date>.xlsx
"""

import os
import re
import json
import glob
import sqlite3
import datetime
import hashlib
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from rapidfuzz import fuzz
from rapidfuzz.process import cdist

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ORDER_DIR   = os.path.join(BASE_DIR, 'data', 'source', 'Ecommerce_Order')
EXPORT_DIR  = os.path.join(BASE_DIR, 'data', 'exports')
DB_PATH     = os.path.join(os.path.dirname(__file__), 'instance', 'inventory.db')
os.makedirs(EXPORT_DIR, exist_ok=True)

TODAY = datetime.date.today().strftime('%Y%m%d')
OUTPUT = os.path.join(EXPORT_DIR, f'ecommerce_mapping_analysis_{TODAY}.xlsx')

# ── Patches ───────────────────────────────────────────────────────────────────
try:
    import openpyxl.worksheet.views as _v
    _orig = _v.Pane.__init__
    def _patch(self, **kw):
        if kw.get('activePane') not in {'bottomRight','bottomLeft','topLeft','topRight'}:
            kw.pop('activePane', None)
        _orig(self, **kw)
    _v.Pane.__init__ = _patch
except: pass

import warnings
warnings.filterwarnings('ignore')


# ── qty_per_sale detection ────────────────────────────────────────────────────
QTY_PATTERNS = [
    re.compile(r'\((\d+)\s*ตัว\)'),       # (50ตัว)
    re.compile(r'\((\d+)\s*ดอก\)'),       # (100ดอก)
    re.compile(r'\((\d+)\s*ชิ้น\)'),      # (10ชิ้น)
    re.compile(r'\((\d+)\s*แพ็ค\)'),      # (1แพ็ค)
    re.compile(r'\((\d+)\s*pcs?\)', re.I),
    re.compile(r'\[(\d+)\s*ตัว\]'),
    re.compile(r'\[(\d+)\s*ดอก\]'),
    re.compile(r'\[(\d+)\s*ชิ้น\]'),
    re.compile(r'(\d+)\s*ตัว/แพ็ค'),
    re.compile(r'(\d+)\s*ตัว\s*$'),
]

def detect_qty_per_sale(name, variation):
    """Auto-detect quantity per sale from name/variation strings."""
    text = f"{name} {variation or ''}"
    for pat in QTY_PATTERNS:
        m = pat.search(text)
        if m:
            try: return int(m.group(1))
            except: pass
    return 1


# ── Parse order files ─────────────────────────────────────────────────────────
def parse_shopee_orders(path):
    """Returns list of (listing_key, item_name, variation, seller_sku, qty, price, date, status)."""
    df = pd.read_excel(path, dtype=str)
    if 'ชื่อสินค้า' not in df.columns:
        return []
    df = df.dropna(subset=['ชื่อสินค้า'])
    out = []
    for _, r in df.iterrows():
        name = str(r['ชื่อสินค้า']).strip()
        var  = str(r.get('ชื่อตัวเลือก','') or '').strip() or None
        sku  = str(r.get('เลขอ้างอิง SKU (SKU Reference No.)','') or '').strip() or None
        qty  = _to_int(r.get('จำนวน')) or 0
        price = _to_float(r.get('ราคาขาย'))
        date = str(r.get('วันที่ทำการสั่งซื้อ','') or '')[:10]
        status = str(r.get('สถานะการสั่งซื้อ','') or '').strip()
        key = hashlib.sha256(f"shopee|{name}|{var or ''}".encode()).hexdigest()[:16]
        out.append({
            'listing_key': key, 'platform': 'shopee', 'item_name': name,
            'variation': var, 'seller_sku': sku, 'qty': qty,
            'price': price, 'date': date, 'status': status,
        })
    return out


def parse_lazada_orders(path):
    df = pd.read_excel(path, dtype=str)
    if 'itemName' not in df.columns:
        return []
    df = df.dropna(subset=['itemName'])
    out = []
    for _, r in df.iterrows():
        name = str(r['itemName']).strip()
        var  = str(r.get('variation','') or '').strip() or None
        sku  = str(r.get('sellerSku','') or '').strip() or None
        # Lazada: each row = 1 item (no qty column visible)
        price = _to_float(r.get('unitPrice'))
        date = str(r.get('createTime','') or '')[:11]
        status = str(r.get('status','') or '').strip()
        key = hashlib.sha256(f"lazada|{name}|{var or ''}".encode()).hexdigest()[:16]
        out.append({
            'listing_key': key, 'platform': 'lazada', 'item_name': name,
            'variation': var, 'seller_sku': sku, 'qty': 1,
            'price': price, 'date': date, 'status': status,
        })
    return out


def _to_int(v):
    try: return int(float(str(v).replace(',','')))
    except: return None

def _to_float(v):
    try: return float(str(v).replace(',',''))
    except: return None


# ── Main analysis ─────────────────────────────────────────────────────────────
def main():
    print('1. Reading order files...')
    all_rows = []
    files = sorted(glob.glob(os.path.join(ORDER_DIR, '*.xlsx')))
    for f in files:
        fname = os.path.basename(f)
        if fname.startswith('.') or 'shop-stats' in fname:
            continue
        if fname.startswith('Order.all.'):
            rows = parse_shopee_orders(f)
            all_rows.extend(rows)
            print(f'  Shopee {fname}: {len(rows)} rows')
        elif fname.startswith('c9566') or 'lazada' in fname.lower():
            rows = parse_lazada_orders(f)
            all_rows.extend(rows)
            print(f'  Lazada {fname}: {len(rows)} rows')

    print(f'\nTotal: {len(all_rows)} order rows')

    # Aggregate per listing
    print('\n2. Aggregating per listing...')
    listings = {}  # listing_key -> aggregated data
    for r in all_rows:
        k = r['listing_key']
        if k not in listings:
            listings[k] = {
                'listing_key': k,
                'platform': r['platform'],
                'item_name': r['item_name'],
                'variation': r['variation'],
                'seller_sku': r['seller_sku'],
                'total_qty': 0,
                'total_revenue': 0.0,
                'orders': 0,
                'min_price': None,
                'max_price': None,
                'min_date': r['date'],
                'max_date': r['date'],
                'cancelled_qty': 0,
            }
        L = listings[k]
        # Skip cancelled
        cancelled = ('ยกเลิก' in r['status']) or (r['status'].lower() in ('canceled','cancelled'))
        if cancelled:
            L['cancelled_qty'] += r['qty']
            continue
        L['total_qty']     += r['qty']
        if r['price']:
            L['total_revenue'] += r['qty'] * r['price']
            L['min_price'] = r['price'] if L['min_price'] is None else min(L['min_price'], r['price'])
            L['max_price'] = r['price'] if L['max_price'] is None else max(L['max_price'], r['price'])
        L['orders'] += 1
        if r['date'] and (not L['min_date'] or r['date'] < L['min_date']):
            L['min_date'] = r['date']
        if r['date'] and (not L['max_date'] or r['date'] > L['max_date']):
            L['max_date'] = r['date']

    print(f'  Unique listings: {len(listings)}')

    # Detect qty_per_sale for each
    for L in listings.values():
        L['qty_per_sale'] = detect_qty_per_sale(L['item_name'], L['variation'])
        L['total_units']  = L['total_qty'] * L['qty_per_sale']

    # ── Load ERP products ──
    print('\n3. Loading ERP products...')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    products = list(conn.execute(
        'SELECT id, sku, product_name, base_sell_price, unit_type FROM products WHERE is_active=1'
    ).fetchall())
    products = [dict(p) for p in products]
    print(f'  {len(products)} active products')

    # ── BSN sales totals per SKU per channel ──
    print('\n4. Loading BSN Z/L sales...')
    bsn_z = defaultdict(lambda: {'qty': 0.0, 'rev': 0.0})  # sku -> totals
    bsn_l = defaultdict(lambda: {'qty': 0.0, 'rev': 0.0})
    rows = conn.execute("""
        SELECT customer_code, p.sku, p.id pid, SUM(qty) q, SUM(net) net
        FROM sales_transactions st
        JOIN products p ON p.id = st.product_id
        WHERE customer_code IN ('Zหน้าร้าน','Lหน้าร้าน')
          AND date_iso < '2026-04-25'
        GROUP BY customer_code, p.sku
    """).fetchall()
    for r in rows:
        target = bsn_z if r['customer_code']=='Zหน้าร้าน' else bsn_l
        target[r['sku']] = {'qty': float(r['q'] or 0), 'rev': float(r['net'] or 0), 'pid': r['pid']}

    # ── Invoice price per product (median per channel + overall) ──
    print('   Loading invoice prices...')
    invoice_price = {}  # pid -> {'z': median_price, 'l': median_price, 'all': median_price}
    rows = conn.execute("""
        SELECT product_id, customer_code, unit_price
        FROM sales_transactions
        WHERE product_id IS NOT NULL AND unit_price > 0
          AND date_iso < '2026-04-25'
    """).fetchall()
    by_pid = defaultdict(lambda: {'z': [], 'l': [], 'all': []})
    for r in rows:
        pid = r['product_id']
        price = float(r['unit_price'])
        by_pid[pid]['all'].append(price)
        if r['customer_code'] == 'Zหน้าร้าน':
            by_pid[pid]['z'].append(price)
        elif r['customer_code'] == 'Lหน้าร้าน':
            by_pid[pid]['l'].append(price)
    def _median(lst):
        if not lst: return None
        s = sorted(lst); n = len(s)
        return s[n//2] if n % 2 else (s[n//2-1] + s[n//2]) / 2
    for pid, d in by_pid.items():
        invoice_price[pid] = {
            'z':   _median(d['z']),
            'l':   _median(d['l']),
            'all': _median(d['all']),
        }

    # ── Fuzzy match ──
    print('\n5. Fuzzy matching listings → ERP SKU...')
    listing_list = list(listings.values())

    def clean(t):
        t = re.sub(r'[\(\[【].*?[\)\]】]', ' ', t or '')
        t = re.sub(r'\b(sendai|golden\s*lion|สิงห์|ม้าลอดห่วง|ของแท้|คุณภาพดี|ฟรี)\b', ' ', t, flags=re.I)
        t = re.sub(r'\s+', ' ', t).strip()
        return t

    corpus = [clean(p['product_name']) for p in products]
    queries = [clean(f"{L['item_name']} {L['variation'] or ''}") for L in listing_list]
    matrix = cdist(queries, corpus, scorer=fuzz.token_set_ratio, workers=-1)
    best_idx = matrix.argmax(axis=1)
    best_score = matrix.max(axis=1)

    for i, L in enumerate(listing_list):
        score = int(best_score[i])
        p = products[best_idx[i]]
        L['suggested_sku']   = p['sku']
        L['suggested_pid']   = p['id']
        L['suggested_name']  = p['product_name']
        L['suggested_unit']  = p['unit_type']
        L['confidence']      = score
        # Pick invoice price: prefer same channel, fallback to overall, then base_sell_price
        ip = invoice_price.get(p['id'], {})
        if L['platform'] == 'shopee':
            erp_price = ip.get('z') or ip.get('all') or p['base_sell_price']
        else:
            erp_price = ip.get('l') or ip.get('all') or p['base_sell_price']
        L['erp_invoice_price'] = erp_price
        # Adjust for qty_per_sale: platform listing price is per pack of qty_per_sale units,
        # so compare (platform_price / qty_per_sale) vs ERP unit price
        if L['min_price'] and erp_price and L['qty_per_sale']:
            per_unit_listing = L['min_price'] / L['qty_per_sale']
            L['price_ratio'] = round(per_unit_listing / erp_price, 2) if erp_price else None
        else:
            L['price_ratio'] = None

    # ── BSN cross-check per suggested SKU ──
    print('\n6. Cross-checking vs BSN Z/L...')
    # Aggregate listings per (platform, suggested_sku)
    sku_totals = defaultdict(lambda: {'shopee_units': 0, 'lazada_units': 0, 'shopee_listings': [], 'lazada_listings': []})
    for L in listing_list:
        s = L['suggested_sku']
        if L['platform'] == 'shopee':
            sku_totals[s]['shopee_units'] += L['total_units']
            sku_totals[s]['shopee_listings'].append(L['listing_key'])
        else:
            sku_totals[s]['lazada_units'] += L['total_units']
            sku_totals[s]['lazada_listings'].append(L['listing_key'])

    # Build cross-check rows
    crosscheck = []
    all_skus = set(sku_totals.keys()) | set(bsn_z.keys()) | set(bsn_l.keys())
    for sku in all_skus:
        p = next((pp for pp in products if pp['sku'] == sku), None)
        if not p:
            continue
        rec = {
            'sku': sku,
            'product_name': p['product_name'],
            'unit_type': p['unit_type'],
            'shopee_listing_units': sku_totals[sku]['shopee_units'],
            'bsn_z_units': bsn_z.get(sku, {}).get('qty', 0),
            'lazada_listing_units': sku_totals[sku]['lazada_units'],
            'bsn_l_units': bsn_l.get(sku, {}).get('qty', 0),
        }
        rec['shopee_diff'] = rec['shopee_listing_units'] - rec['bsn_z_units']
        rec['lazada_diff'] = rec['lazada_listing_units'] - rec['bsn_l_units']
        crosscheck.append(rec)

    crosscheck.sort(key=lambda r: abs(r['shopee_diff']) + abs(r['lazada_diff']), reverse=True)

    # ── Write Excel ──
    print(f'\n7. Writing {OUTPUT}')
    wb = openpyxl.Workbook()
    write_listings_sheet(wb, listing_list, bsn_z, bsn_l)
    write_crosscheck_sheet(wb, crosscheck)
    write_summary_sheet(wb, listing_list, crosscheck)

    wb.save(OUTPUT)
    print(f'\nDone! Output: {OUTPUT}')
    print(f'  Listings: {len(listing_list)}')
    print(f'  Cross-check rows: {len(crosscheck)}')


# ── Excel writers ─────────────────────────────────────────────────────────────
def write_listings_sheet(wb, listings, bsn_z, bsn_l):
    ws = wb.active
    ws.title = 'Listing Mapping'

    cols = [
        'platform', 'ชื่อสินค้า (Platform)', 'ตัวเลือก', 'seller_sku',
        'จำนวนออเดอร์', 'รวม qty (ชิ้น)', 'qty_per_sale (ตรวจจับ)', 'รวม units (ERP)',
        'ยกเลิก qty', 'ราคาต่ำสุด', 'ราคาสูงสุด', 'รายได้รวม',
        'วันแรก', 'วันสุดท้าย',
        'Suggested SKU', 'Suggested Name (ERP)', 'หน่วย ERP', 'ราคา ERP (invoice mid)', 'ratio (Plat/unit ÷ ERP)', 'confidence%',
        'BSN Z qty', 'Z diff (units - BSN)',
        'BSN L qty', 'L diff (units - BSN)',
    ]
    fill_hdr  = PatternFill('solid', start_color='1F4E78')
    font_hdr  = Font(bold=True, color='FFFFFF', size=10)
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(1, ci, c)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 38

    fill_hi  = PatternFill('solid', start_color='E8F5E9')
    fill_mid = PatternFill('solid', start_color='FFF9C4')
    fill_lo  = PatternFill('solid', start_color='FFCCBC')

    for ri, L in enumerate(sorted(listings, key=lambda x: (x['platform'], -x['total_units'])), 2):
        z = bsn_z.get(L['suggested_sku'], {}).get('qty', 0)
        l_ = bsn_l.get(L['suggested_sku'], {}).get('qty', 0)

        if L['platform'] == 'shopee':
            z_diff = L['total_units'] - z
            l_diff = ''
        else:
            z_diff = ''
            l_diff = L['total_units'] - l_

        vals = [
            L['platform'], L['item_name'], L['variation'] or '', L['seller_sku'] or '',
            L['orders'], L['total_qty'], L['qty_per_sale'], L['total_units'],
            L['cancelled_qty'], L['min_price'], L['max_price'], round(L['total_revenue'], 2),
            L['min_date'], L['max_date'],
            L['suggested_sku'], L['suggested_name'], L['suggested_unit'],
            L['erp_invoice_price'], L['price_ratio'] if L['price_ratio'] else '',
            L['confidence'],
            z if L['platform']=='shopee' else '',
            z_diff,
            l_ if L['platform']=='lazada' else '',
            l_diff,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(size=9)

        # Confidence color
        cell_conf = ws.cell(ri, cols.index('confidence%')+1)
        if L['confidence'] >= 80:
            cell_conf.fill = fill_hi
        elif L['confidence'] >= 60:
            cell_conf.fill = fill_mid
        else:
            cell_conf.fill = fill_lo

    widths = [9, 50, 25, 15, 10, 12, 14, 14, 10, 11, 11, 13, 11, 11,
              12, 38, 9, 11, 11, 12, 10, 14, 10, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'


def write_crosscheck_sheet(wb, rows):
    ws = wb.create_sheet('Cross-Check by SKU')
    cols = ['SKU','ชื่อสินค้า','หน่วย',
            'Shopee listings (units)','BSN Z qty','Diff Shopee',
            'Lazada listings (units)','BSN L qty','Diff Lazada',
            'หมายเหตุ']
    fill_hdr  = PatternFill('solid', start_color='1F4E78')
    font_hdr  = Font(bold=True, color='FFFFFF', size=10)
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(1, ci, c)
        cell.fill = fill_hdr; cell.font = font_hdr
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    for ri, r in enumerate(rows, 2):
        notes = []
        if abs(r['shopee_diff']) > 5: notes.append(f'Shopee ต่าง {r["shopee_diff"]:+.0f}')
        if abs(r['lazada_diff']) > 5: notes.append(f'Lazada ต่าง {r["lazada_diff"]:+.0f}')
        vals = [r['sku'], r['product_name'], r['unit_type'],
                r['shopee_listing_units'], r['bsn_z_units'], r['shopee_diff'],
                r['lazada_listing_units'], r['bsn_l_units'], r['lazada_diff'],
                ' | '.join(notes)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(size=9)
            if ci in (6, 9) and isinstance(v, (int, float)):
                if abs(v) > 5:
                    c.fill = PatternFill('solid', start_color='FFCCBC')
                elif abs(v) > 0:
                    c.fill = PatternFill('solid', start_color='FFF9C4')
                else:
                    c.fill = PatternFill('solid', start_color='E8F5E9')

    widths = [10, 50, 9, 15, 11, 11, 15, 11, 11, 30]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'


def write_summary_sheet(wb, listings, crosscheck):
    ws = wb.create_sheet('Summary', 0)
    ws['A1'] = 'สรุปการวิเคราะห์ Mapping Shopee/Lazada'
    ws['A1'].font = Font(bold=True, size=14)

    shopee = [L for L in listings if L['platform']=='shopee']
    lazada = [L for L in listings if L['platform']=='lazada']

    rows = [
        ('', ''),
        ('Listing Shopee', len(shopee)),
        ('Listing Lazada', len(lazada)),
        ('รวม listing', len(listings)),
        ('', ''),
        ('Shopee total orders', sum(L['orders'] for L in shopee)),
        ('Shopee total qty (raw)', sum(L['total_qty'] for L in shopee)),
        ('Shopee total units (×qty_per_sale)', sum(L['total_units'] for L in shopee)),
        ('Shopee total revenue', round(sum(L['total_revenue'] for L in shopee), 2)),
        ('', ''),
        ('Lazada total orders', sum(L['orders'] for L in lazada)),
        ('Lazada total qty', sum(L['total_qty'] for L in lazada)),
        ('Lazada total units', sum(L['total_units'] for L in lazada)),
        ('Lazada total revenue', round(sum(L['total_revenue'] for L in lazada), 2)),
        ('', ''),
        ('confidence ≥80%', sum(1 for L in listings if L['confidence']>=80)),
        ('confidence 60-79%', sum(1 for L in listings if 60<=L['confidence']<80)),
        ('confidence <60%', sum(1 for L in listings if L['confidence']<60)),
        ('', ''),
        ('Cross-check rows', len(crosscheck)),
        ('Diff Shopee >5', sum(1 for r in crosscheck if abs(r['shopee_diff'])>5)),
        ('Diff Lazada >5', sum(1 for r in crosscheck if abs(r['lazada_diff'])>5)),
    ]
    for ri, (k, v) in enumerate(rows, 3):
        ws.cell(ri, 1, k).font = Font(size=11, bold=bool(k and 'รวม' in k))
        ws.cell(ri, 2, v).font = Font(size=11)

    ws.cell(30, 1, 'คำอธิบาย').font = Font(bold=True, size=12)
    notes = [
        'Listing Mapping: หนึ่งแถวต่อ 1 listing ที่ไม่ซ้ำ (ชื่อ+ตัวเลือก)',
        'qty_per_sale: ระบบตรวจจับจาก variation/ชื่อ เช่น "(50ตัว)" → 50',
        'รวม units (ERP) = รวม qty × qty_per_sale → จำนวนหน่วยจริงใน ERP',
        'BSN Z qty / L qty: ยอดขายในระบบที่ลูกค้าคือ Zหน้าร้าน (Shopee) / Lหน้าร้าน (Lazada)',
        'Diff: รวม units - BSN qty → ถ้าเป็น 0 = ตรงกัน, ติดลบ = BSN เกิน, บวก = BSN ขาด',
        'confidence: ระบบจับคู่ชื่อ — สีเขียว ≥80% / เหลือง 60-79% / แดง <60%',
        'ratio: ราคา platform/ราคา ERP base — ใกล้ 1.0 = น่าจะถูก',
    ]
    for i, n in enumerate(notes, 31):
        ws.cell(i, 1, '• ' + n).font = Font(size=10)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20


if __name__ == '__main__':
    main()
