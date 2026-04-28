"""
parse_platform.py — Import/Export Shopee & Lazada xlsx files

Shopee Mass Update format:
  Row 0: metadata
  Row 1: (filler)
  Row 2: Thai column headers  ← header row
  Row 3-4: instruction rows
  Row 5+: actual data

Lazada Price/Stock Export format:
  Row 0-2: instruction rows
  Row 3+:  actual data (header is the pandas default from read_excel)
"""

import io
import json
import re
import warnings

# ── Patch openpyxl bug (Shopee xlsx has invalid activePane value) ─────────────
try:
    import openpyxl.worksheet.views as _opxl_views
    _orig_pane_init = _opxl_views.Pane.__init__

    def _patched_pane_init(self, **kwargs):
        valid = {'bottomRight', 'bottomLeft', 'topLeft', 'topRight'}
        if kwargs.get('activePane') not in valid:
            kwargs.pop('activePane', None)
        _orig_pane_init(self, **kwargs)

    _opxl_views.Pane.__init__ = _patched_pane_init
except Exception:
    pass


import pandas as pd

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# ── Shopee ────────────────────────────────────────────────────────────────────

SHOPEE_METADATA_ROWS = 2   # rows before Thai header (rows 0-1)
SHOPEE_SKIP_AFTER_HDR = 3  # instruction rows after header to skip (rows 3-5)


def parse_shopee(file_obj):
    """
    Parse Shopee Mass Update xlsx.
    Returns list of dicts with keys:
      product_id_str, product_name, variation_id, variation_name,
      parent_sku, seller_sku, price, stock, raw_json
    """
    df = pd.read_excel(file_obj, header=SHOPEE_METADATA_ROWS, dtype=str)
    # Drop instruction rows (first SHOPEE_SKIP_AFTER_HDR rows after header)
    df = df.iloc[SHOPEE_SKIP_AFTER_HDR:].reset_index(drop=True)
    # Keep only rows with numeric product ID
    df = df[df['รหัสสินค้า'].str.match(r'^\d+$', na=False)].copy()

    records = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else v) for k, v in row.items()}
        records.append({
            'product_id_str':  raw.get('รหัสสินค้า'),
            'product_name':    raw.get('ชื่อสินค้า') or '',
            'variation_id':    raw.get('รหัสตัวเลือกสินค้า'),
            'variation_name':  raw.get('ชื่อตัวเลือกสินค้า'),
            'parent_sku':      raw.get('Parent SKU'),
            'seller_sku':      raw.get('เลข SKU'),
            'price':           _to_float(raw.get('ราคา')),
            'special_price':   None,
            'stock':           _to_int(raw.get('คลัง')),
            'raw_json':        json.dumps(raw, ensure_ascii=False),
        })
    return records


def export_shopee(rows):
    """
    Generate Shopee Mass Update xlsx (BytesIO) from list of platform_skus rows.
    Preserves original raw_json columns; only price & stock are from DB.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    COLS = [
        'รหัสสินค้า', 'ชื่อสินค้า', 'รหัสตัวเลือกสินค้า', 'ชื่อตัวเลือกสินค้า',
        'Parent SKU', 'เลข SKU', 'ราคา', 'GTIN', 'คลัง', 'จำนวนการซื้อขั้นต่ำ',
        'จำนวนสินค้าสูงสุดที่ซื้อได้',
        'จำนวนสินค้าสูงสุดที่ซื้อได้ - เวลาเริ่มต้น',
        'จำนวนสินค้าสูงสุดที่ซื้อได้ - ระยะเวลา (วัน)',
        'จำนวนสินค้าสูงสุดที่ซื้อได้ - วันที่สิ้นสุด',
        'เหตุผล',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row (row 1 = Shopee's row-2 style, orange background)
    orange = PatternFill('solid', start_color='FF6633')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = orange
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for ri, row in enumerate(rows, 2):
        raw = json.loads(row['raw_json']) if row['raw_json'] else {}
        # Override price & stock with current DB values
        raw['ราคา']  = row['price']  if row['price']  is not None else raw.get('ราคา')
        raw['คลัง'] = row['stock'] if row['stock'] is not None else raw.get('คลัง')
        for ci, col in enumerate(COLS, 1):
            ws.cell(row=ri, column=ci, value=raw.get(col))

    # Column widths
    widths = [15, 40, 22, 25, 15, 15, 10, 10, 8, 18, 20, 28, 24, 26, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Lazada ────────────────────────────────────────────────────────────────────

LAZADA_SKIP_ROWS  = 3   # rows 0-2 are instructions; row 3 = first data


def parse_lazada(file_obj):
    """
    Parse Lazada Price/Stock Export xlsx.
    Returns list of dicts.
    """
    df = pd.read_excel(file_obj, dtype=str)
    df = df.iloc[LAZADA_SKIP_ROWS:].reset_index(drop=True)
    # Keep rows with numeric Product ID
    df = df[df['Product ID'].str.match(r'^\d+$', na=False)].copy()

    # The stock column header is the store name ("บุญสวัสดิ์นำชัย")
    stock_col = _find_stock_col(df)

    records = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else v) for k, v in row.items()}
        records.append({
            'product_id_str': raw.get('Product ID'),
            'product_name':   raw.get('ชื่อสินค้า') or '',
            'variation_id':   raw.get('ร้าน sku'),
            'variation_name': raw.get('Variations Combo'),
            'parent_sku':     None,
            'seller_sku':     raw.get('SellerSKU'),
            'price':          _to_float(raw.get('ราคา')),
            'special_price':  _to_float(raw.get('SpecialPrice')),
            'stock':          _to_int(raw.get(stock_col)),
            'raw_json':       json.dumps(raw, ensure_ascii=False),
        })
    return records


def _find_stock_col(df):
    """Find the stock column — it's the store-name column between SellerSKU and Variations Combo."""
    # Columns that are NOT these known ones, but come after 'ราคา' → is the stock col
    known = {'Product ID', 'catId', 'ชื่อสินค้า', 'currencyCode', 'sku.skuId',
             'status', 'ร้าน sku', 'SpecialPrice', 'SpecialPrice Start',
             'SpecialPrice End', 'ราคา', 'SellerSKU', 'Variations Combo',
             'tr(s-wb-product@md5key)'}
    for col in df.columns:
        if col not in known:
            return col
    return 'บุญสวัสดิ์นำชัย'  # fallback


def export_lazada(rows):
    """
    Generate Lazada Price/Stock update xlsx (BytesIO).
    Only includes columns Lazada needs for batch price/stock update.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    COLS = ['ร้าน sku', 'ราคา', 'SpecialPrice', 'คลัง (ร้าน)']

    wb = openpyxl.Workbook()
    ws = wb.active

    blue = PatternFill('solid', start_color='003087')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = blue
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, row in enumerate(rows, 2):
        raw = json.loads(row['raw_json']) if row['raw_json'] else {}
        stock_col = _find_stock_col_raw(raw)
        ws.cell(row=ri, column=1, value=row['variation_id'])
        ws.cell(row=ri, column=2, value=row['price']  if row['price']  is not None else raw.get('ราคา'))
        ws.cell(row=ri, column=3, value=row['special_price'] if row['special_price'] is not None else raw.get('SpecialPrice'))
        ws.cell(row=ri, column=4, value=row['stock'] if row['stock'] is not None else raw.get(stock_col))

    for ci, w in enumerate([25, 12, 14, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Add a full-data sheet
    ws_full = wb.create_sheet('ข้อมูลทั้งหมด')
    if rows:
        all_raw = [json.loads(r['raw_json']) for r in rows if r['raw_json']]
        if all_raw:
            full_cols = list(all_raw[0].keys())
            for ci, col in enumerate(full_cols, 1):
                ws_full.cell(row=1, column=ci, value=col).font = Font(bold=True)
            for ri, raw in enumerate(all_raw, 2):
                price_key = 'ราคา'
                stock_key = _find_stock_col_raw(raw)
                # Override with DB values
                matching = [r for r in rows if json.loads(r['raw_json']).get('ร้าน sku') == raw.get('ร้าน sku')]
                if matching:
                    r = matching[0]
                    raw[price_key] = r['price'] if r['price'] is not None else raw.get(price_key)
                    raw['SpecialPrice'] = r['special_price'] if r['special_price'] is not None else raw.get('SpecialPrice')
                    if stock_key:
                        raw[stock_key] = r['stock'] if r['stock'] is not None else raw.get(stock_key)
                for ci, col in enumerate(full_cols, 1):
                    ws_full.cell(row=ri, column=ci, value=raw.get(col))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _find_stock_col_raw(raw_dict):
    known = {'Product ID', 'catId', 'ชื่อสินค้า', 'currencyCode', 'sku.skuId',
             'status', 'ร้าน sku', 'SpecialPrice', 'SpecialPrice Start',
             'SpecialPrice End', 'ราคา', 'SellerSKU', 'Variations Combo',
             'tr(s-wb-product@md5key)'}
    for k in raw_dict:
        if k not in known:
            return k
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_float(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return None


def _to_int(val):
    if val is None:
        return None
    try:
        return int(float(str(val).replace(',', '')))
    except (ValueError, TypeError):
        return None


# ── Platform Mapping Export/Import ───────────────────────────────────────────

MAPPING_COLS = [
    # Read-only info (col 1-10)
    'platform_sku_id', 'platform', 'รหัสสินค้า (platform)',
    'ชื่อสินค้า (platform)', 'variation_id', 'ชื่อ variation', 'seller_sku',
    'ราคา (platform)', 'ราคาพิเศษ (platform, Lazada)', 'คลัง (platform)',
    # AI suggestion (col 11-13, pre-filled but editable)
    'internal_sku', 'ชื่อสินค้า (ระบบ) — อ่านอย่างเดียว', 'confidence_%',
    # User adjusts (col 14)
    'qty_per_sale',
    # Hint (col 15)
    'คำอธิบาย qty_per_sale',
]

_READONLY_COLS = 10   # cols 1-10 = read-only platform data
_SUGGESTION_COLS = 3  # cols 11-13 = AI suggestions (editable but pre-filled)


def export_mapping(rows, suggestions=None):
    """
    Generate mapping xlsx for user to fill in.
    rows:        list of sqlite3.Row from get_platform_mapping_data()
    suggestions: dict {sku_id -> {suggested_sku, suggested_name, confidence}}
    Returns BytesIO.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    suggestions = suggestions or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mapping'

    # ── Header ──
    fill_info  = PatternFill('solid', start_color='FFF3CD')  # yellow = platform info
    fill_ai    = PatternFill('solid', start_color='D1ECF1')  # blue   = AI suggestion
    fill_edit  = PatternFill('solid', start_color='D4EDDA')  # green  = user edits
    fill_hint  = PatternFill('solid', start_color='F8F9FA')  # grey   = hint
    hdr_font   = Font(bold=True, size=9)

    for ci, col in enumerate(MAPPING_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if ci <= _READONLY_COLS:
            cell.fill = fill_info
        elif ci <= _READONLY_COLS + _SUGGESTION_COLS:
            cell.fill = fill_ai
        elif ci < len(MAPPING_COLS):
            cell.fill = fill_edit
        else:
            cell.fill = fill_hint
    ws.row_dimensions[1].height = 40

    # Row fills by confidence band
    fill_hi   = PatternFill('solid', start_color='F0FFF4')  # ≥80%: pale green
    fill_mid  = PatternFill('solid', start_color='FFFDE7')  # 60-79%: pale yellow
    fill_lo   = PatternFill('solid', start_color='FFF3E0')  # 40-59%: pale orange
    fill_vlo  = PatternFill('solid', start_color='FFF5F5')  # <40%: pale red

    # ── Data rows ──
    for ri, r in enumerate(rows, 2):
        r = dict(r)
        sg = suggestions.get(r['id'], {})
        conf = sg.get('confidence', 0)

        # Determine row background by confidence
        if conf == 100:
            row_fill = PatternFill('solid', start_color='E8F5E9')  # already mapped
        elif conf >= 80:
            row_fill = fill_hi
        elif conf >= 60:
            row_fill = fill_mid
        elif conf >= 40:
            row_fill = fill_lo
        else:
            row_fill = fill_vlo

        # Use existing mapping first, fall back to suggestion
        int_sku  = r.get('internal_sku') or (sg.get('suggested_sku', '') if sg else '')
        int_name = r.get('internal_product_name') or (sg.get('suggested_name', '') if sg else '')
        conf_val = 100 if r.get('internal_sku') else (sg.get('confidence', '') if sg else '')

        vals = [
            r['id'],
            r['platform'],
            r['product_id_str'],
            r['product_name'],
            r['variation_id'],
            r['variation_name'] or '',
            r['seller_sku'] or '',
            r['price'],
            r['special_price'],
            r['stock'],
            # AI / existing suggestion (editable)
            int_sku,
            int_name,
            conf_val,
            # User edits
            r['qty_per_sale'] or 1,
            # Hint
            'จำนวนหน่วยในระบบที่ลดเมื่อขาย 1 ชิ้น (เช่น 0.5 กก., 6 ม้วน, 100 ตัว)',
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
            if ci <= _READONLY_COLS:
                cell.fill = row_fill
            elif ci == 12:  # ชื่อสินค้าในระบบ = read-only suggestion
                cell.fill = PatternFill('solid', start_color='EBF5FB')
                cell.font = Font(size=9, italic=True, color='2471A3')
            elif ci == 13:  # confidence%
                # Color the confidence cell
                if conf_val == '' or conf_val == 0:
                    cell.fill = PatternFill('solid', start_color='F5F5F5')
                elif conf_val == 100:
                    cell.fill = PatternFill('solid', start_color='C8E6C9')
                    cell.font = Font(size=9, bold=True, color='1B5E20')
                elif conf_val >= 80:
                    cell.fill = PatternFill('solid', start_color='DCEDC8')
                    cell.font = Font(size=9, bold=True, color='33691E')
                elif conf_val >= 60:
                    cell.fill = PatternFill('solid', start_color='FFF9C4')
                    cell.font = Font(size=9, bold=True, color='F57F17')
                else:
                    cell.fill = PatternFill('solid', start_color='FFCCBC')
                    cell.font = Font(size=9, bold=True, color='BF360C')
            elif ci == len(MAPPING_COLS):  # hint
                cell.fill = PatternFill('solid', start_color='F8F9FA')
                cell.font = Font(size=8, italic=True, color='999999')

    # ── Column widths ──
    widths = [10, 9, 16, 42, 26, 22, 14, 9, 18, 9, 13, 36, 12, 12, 52]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A2'

    # ── Legend sheet ──
    ws2 = wb.create_sheet('คำอธิบาย')
    legends = [
        ('สีเขียวเข้ม (100%)', 'ผูกกับสินค้าในระบบแล้ว'),
        ('สีเขียวอ่อน (80-99%)', 'AI มั่นใจสูง — น่าจะถูกต้อง'),
        ('สีเหลือง (60-79%)', 'AI คิดว่าน่าจะใช่ — ตรวจสอบด้วย'),
        ('สีส้ม (40-59%)', 'AI ไม่แน่ใจ — ต้องแก้ไขเอง'),
        ('สีแดงอ่อน (<40%)', 'AI ไม่มั่นใจ — กรุณากรอก internal_sku เอง'),
        ('', ''),
        ('internal_sku', 'ใส่ตัวเลข SKU ของสินค้าในระบบ ERP'),
        ('qty_per_sale', 'จำนวนหน่วยในระบบที่ลดเมื่อขาย 1 ชิ้นบน platform'),
        ('ตัวอย่าง qty_per_sale', 'ปุ๊ก #7 (500 กรัม) = 0.5 | สายเอ็น 6 ม้วน = 6 | ลูกรีเวท 100 ตัว = 100'),
    ]
    ws2['A1'] = 'สัญลักษณ์สี'
    ws2['B1'] = 'ความหมาย'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    for ri2, (k, v) in enumerate(legends, 2):
        ws2.cell(row=ri2, column=1, value=k)
        ws2.cell(row=ri2, column=2, value=v)
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_mapping(file_obj):
    """
    Parse a filled mapping xlsx.
    Returns list of dicts: {platform_sku_id, internal_sku, qty_per_sale}
    """
    df = pd.read_excel(file_obj, dtype=str)
    results = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else str(v).strip()) for k, v in row.items()}
        sku_id = _to_int(raw.get('platform_sku_id'))
        if not sku_id:
            continue
        results.append({
            'platform_sku_id': sku_id,
            'internal_sku':    raw.get('internal_sku') or None,
            'qty_per_sale':    _to_float(raw.get('qty_per_sale')) or 1.0,
        })
    return results


# ── Order File Parsers (for Listing Mapping) ─────────────────────────────────

def parse_shopee_orders(file_obj):
    """
    Parse Shopee All-Orders xlsx (exported from My Orders).
    Extracts unique (item_name, variation) combinations.
    Returns list of dicts for ecommerce_listings.
    """
    import hashlib
    df = pd.read_excel(file_obj, dtype=str)
    if 'ชื่อสินค้า' not in df.columns:
        raise ValueError('ไม่พบคอลัมน์ ชื่อสินค้า — ตรวจสอบว่าใช้ไฟล์คำสั่งซื้อ Shopee')

    df = df.dropna(subset=['ชื่อสินค้า'])
    seen = {}
    for _, row in df.iterrows():
        name = str(row['ชื่อสินค้า']).strip()
        var  = str(row.get('ชื่อตัวเลือก', '') or '').strip() or None
        key  = hashlib.sha256(f"shopee|{name}|{var or ''}".encode()).hexdigest()[:16]
        if key in seen:
            continue
        price      = _to_float(row.get('ราคาขาย'))
        seller_sku = str(row.get('เลขอ้างอิง SKU (SKU Reference No.)', '') or '').strip() or None
        seen[key] = {
            'platform':     'shopee',
            'item_name':    name,
            'variation':    var,
            'seller_sku':   seller_sku,
            'listing_key':  key,
            'sample_price': price,
        }
    return list(seen.values())


_LAZADA_ATTR_PREFIX_RE = re.compile(
    r'^(สี|โทนสี|Color Family|Color family|Color|Variation\d*|Item|ขนาด|จำนวน|'
    r'Power Tools Battery Voltage|สินค้า)\s*:\s*',
    re.IGNORECASE,
)


def parse_lazada_orders(file_obj):
    """
    Parse Lazada Orders xlsx (exported from Seller Center).
    Extracts unique (itemName, variation) combinations.
    Returns list of dicts for ecommerce_listings.

    Lazada has changed the variation attribute label over time
    (โทนสี: → สี: → Color Family: → Color family:). We strip these prefixes
    before hashing so the same option doesn't create duplicate listings
    across years.
    """
    import hashlib
    df = pd.read_excel(file_obj, dtype=str)
    if 'itemName' not in df.columns:
        raise ValueError('ไม่พบคอลัมน์ itemName — ตรวจสอบว่าใช้ไฟล์คำสั่งซื้อ Lazada')

    df = df.dropna(subset=['itemName'])
    seen = {}
    for _, row in df.iterrows():
        name = str(row['itemName']).strip()
        var  = str(row.get('variation', '') or '').strip() or None
        if var:
            var = _LAZADA_ATTR_PREFIX_RE.sub('', var).strip() or None
        key  = hashlib.sha256(f"lazada|{name}|{var or ''}".encode()).hexdigest()[:16]
        if key in seen:
            continue
        price      = _to_float(row.get('unitPrice'))
        seller_sku = str(row.get('sellerSku', '') or '').strip() or None
        seen[key] = {
            'platform':     'lazada',
            'item_name':    name,
            'variation':    var,
            'seller_sku':   seller_sku,
            'listing_key':  key,
            'sample_price': price,
        }
    return list(seen.values())


# ── Listing Mapping Export/Import ─────────────────────────────────────────────

LISTING_MAPPING_COLS = [
    'listing_id', 'platform', 'ชื่อสินค้า (platform)', 'ตัวเลือก (platform)',
    'seller_sku', 'ราคาตัวอย่าง (฿)',
    'internal_sku', 'ชื่อสินค้า (ERP) — อ่านอย่างเดียว', 'confidence_%',
    'qty_per_sale', 'คำอธิบาย qty_per_sale',
]


def export_listing_mapping(rows, suggestions=None, unmatched_only=False):
    """
    Generate mapping xlsx for user to fill in internal_sku.
    rows: list of ecommerce_listings rows (dicts or sqlite3.Row)
    suggestions: dict {listing_id -> {suggested_sku, suggested_name, confidence}}
    unmatched_only: True → only export rows without product_id
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    suggestions = suggestions or {}
    rows = [dict(r) for r in rows]
    if unmatched_only:
        rows = [r for r in rows if not r.get('product_id')]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mapping'

    fill_info = PatternFill('solid', start_color='FFF3CD')
    fill_edit = PatternFill('solid', start_color='D4EDDA')
    fill_ai   = PatternFill('solid', start_color='D1ECF1')
    hdr_font  = Font(bold=True, size=9)

    for ci, col in enumerate(LISTING_MAPPING_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if col in ('internal_sku',):
            cell.fill = fill_edit
        elif 'ERP' in col or 'confidence' in col:
            cell.fill = fill_ai
        else:
            cell.fill = fill_info
    ws.row_dimensions[1].height = 36

    fill_mapped = PatternFill('solid', start_color='E8F5E9')
    fill_hi     = PatternFill('solid', start_color='F0FFF4')
    fill_mid    = PatternFill('solid', start_color='FFFDE7')
    fill_lo     = PatternFill('solid', start_color='FFF3E0')
    fill_vlo    = PatternFill('solid', start_color='FFF5F5')

    for ri, r in enumerate(rows, 2):
        sg   = suggestions.get(r['id'], {})
        conf = sg.get('confidence', 0)

        int_sku  = r.get('sku') or (sg.get('suggested_sku', '') if sg else '')
        int_name = r.get('product_name') or (sg.get('suggested_name', '') if sg else '')
        conf_val = 100 if r.get('product_id') else (sg.get('confidence', '') if sg else '')

        if r.get('product_id'):
            row_fill = fill_mapped
        elif conf >= 80:
            row_fill = fill_hi
        elif conf >= 60:
            row_fill = fill_mid
        elif conf >= 40:
            row_fill = fill_lo
        else:
            row_fill = fill_vlo

        qty_per_sale = r.get('qty_per_sale') or 1

        vals = [
            r['id'], r['platform'], r['item_name'], r.get('variation') or '',
            r.get('seller_sku') or '', r.get('sample_price'),
            int_sku, int_name, conf_val,
            qty_per_sale,
            'จำนวนหน่วยในระบบที่ลดเมื่อขาย 1 ชิ้น เช่น ลูกรีเวท 25 ดอก/ถุง → ใส่ 25',
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
            col_name = LISTING_MAPPING_COLS[ci - 1]
            if col_name in ('internal_sku', 'qty_per_sale'):
                cell.fill = fill_edit
            elif 'ERP' in col_name:
                cell.fill = PatternFill('solid', start_color='EBF5FB')
                cell.font = Font(size=9, italic=True, color='2471A3')
            elif col_name == 'confidence_%':
                if conf_val == '' or conf_val == 0:
                    cell.fill = PatternFill('solid', start_color='F5F5F5')
                elif conf_val == 100:
                    cell.fill = PatternFill('solid', start_color='C8E6C9')
                    cell.font = Font(size=9, bold=True, color='1B5E20')
                elif conf_val >= 80:
                    cell.fill = PatternFill('solid', start_color='DCEDC8')
                    cell.font = Font(size=9, bold=True, color='33691E')
                elif conf_val >= 60:
                    cell.fill = PatternFill('solid', start_color='FFF9C4')
                    cell.font = Font(size=9, bold=True, color='F57F17')
                else:
                    cell.fill = PatternFill('solid', start_color='FFCCBC')
                    cell.font = Font(size=9, bold=True, color='BF360C')
            elif col_name == 'คำอธิบาย qty_per_sale':
                cell.fill = PatternFill('solid', start_color='F8F9FA')
                cell.font = Font(size=8, italic=True, color='999999')
            else:
                cell.fill = row_fill

    widths = [10, 9, 50, 28, 16, 14, 13, 40, 12, 12, 55]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_listing_mapping(file_obj):
    """
    Parse filled listing mapping xlsx.
    Returns list of dicts: {listing_id, internal_sku, qty_per_sale}
    """
    df = pd.read_excel(file_obj, dtype=str)
    results = []
    for _, row in df.iterrows():
        raw = {k: (None if pd.isna(v) else str(v).strip()) for k, v in row.items()}
        lid = _to_int(raw.get('listing_id'))
        if not lid:
            continue
        results.append({
            'listing_id':    lid,
            'internal_sku':  raw.get('internal_sku') or None,
            'qty_per_sale':  _to_float(raw.get('qty_per_sale')) or 1.0,
        })
    return results
